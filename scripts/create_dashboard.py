import os
import subprocess
import sys
import shutil
from datetime import datetime

# Auto-install openpyxl falls nicht vorhanden
try:
    import openpyxl
except ImportError:
    print("Installiere openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl

from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()
wb.remove(wb.active)

# =========================================================
# BLATT 1: IST-DATEN
# =========================================================
ws1 = wb.create_sheet('1_Ist_Daten')
data1 = [
    ['Feld', 'Wert', 'Einheit', 'Hinweis (Aus letzter Excel/Berichten)'],
    ['Kassenbestand (Start Periode)', 28.05, 'MEUR', 'Aus Bilanz Aktiva'],
    ['Forderungen aus L&L (Vorperiode)', 31.36, 'MEUR', 'Bilanz Aktiva - wird zu Einzahlungen'],
    ['Eigenkapital', 35.25, 'MEUR', 'Aus Bilanz Passiva'],
    ['Gewinnvortrag', 5.27, 'MEUR', 'Maximal mögliche Dividende'],
    ['Steuerlicher Verlustvortrag', 0.00, 'MEUR', 'Vorjahresverlust (vermindert Steuern)'],
    ['Pensionsrückstellungen', 15.95, 'MEUR', 'Passiva'],
    ['Bestehende Kurzfristkredite (Alt)', 12.0, 'MEUR', 'Werden auto-getilgt (fließen zwingend ab!)'],
    ['Alt-Überziehungskredit', 0.00, 'MEUR', 'Aus Vorperiode, falls angefallen'],
    ['Bestehende Langfristkredite (Alt)', 20.0, 'MEUR', 'Nur für Plan-Zinsen wichtig'],
    ['Wertpapiere (Vorperiode)', 0.00, 'MEUR', 'Werden auto-zurückgezahlt'],
    ['Zinssatz Wertpapiere (%)', 0.04, '%', 'Aus Wirtschaftsnachrichten'],
    ['Geplante E-Steuer (%)', 0.45, '%', 'Im Handbuch S.33 klar auf 45% festgelegt'],
    ['Zinssatz Langfristig (%)', 0.07, '%', 'Bsp: 7% aus Handbuch S.31 / Per.0'],
    ['Zinssatz Kurzfristig (%)', 0.08, '%', 'Bsp: 8% aus Berichten / Per.0'],
    ['Zinssatz Überziehen (%)', 0.13, '%', '13% Strafzins (Handbuch)'],
    ['Aktuelles Rating', 'BBB', '', 'Aktuelles Rating für Zinsanpassung'],
    ['Rating-Zinsänderung (%)', -0.01, '%', 'Z.B. BBB = -1% (Handbuch S.32)']
]
for row in data1: ws1.append(row)

# =========================================================
# BLATT 2: TEAM-INPUTS
# =========================================================
ws2 = wb.create_sheet('2_Team_Inputs')

data2_base = [
    ['Feld', 'Wert', 'Einheit', 'Hinweis (Vom Team erfragen)'],
    ['Geplante Absatzmenge', 45000, 'Stück', 'Planzahl Vertrieb'],
    ['Geplanter Preis', 3000, 'EUR', 'Planzahl Vertrieb'],
    ['Geplante Produktionsmenge', 45000, 'Stück', 'Für Bestandsveränderung'],
    ['--- PERSONAL WESEN (MATRIX) ---', 'Einkauf', 'Verwaltung', 'Fertigung', 'F&E', 'Vertrieb', 'Summe'],
    ['Personalanfangsbestand', 20, 214, 857, 40, 100, "=SUM(B6:F6)"],
    ['+ Einstellungen', 0, 0, 50, 0, 0, "=SUM(B7:F7)"],
    ['- Entlassungen', 0, 0, 0, 0, 0, "=SUM(B8:F8)"],
    ['- Fluktuation', 0, 0, 0, 0, 0, "=SUM(B9:F9)"],
    ['= Personalendbestand', "=B6+B7-B8-B9", "=C6+C7-C8-C9", "=D6+D7-D8-D9", "=E6+E7-E8-E9", "=F6+F7-F8-F9", "=SUM(B10:F10)"],
    ['Kosten je Mitarbeiter (EUR)', 31000, 29000, 31000, 45000, 41000, ""],
    ['Kosten je Einstellung (EUR)', 12500, 12500, 12500, 12500, 12500, ""],
    ['Kosten je Entlassung (EUR)', 10000, 10000, 10000, 10000, 10000, ""],
    ['Training je Mitarbeiter (EUR)', 1600, 1600, 1600, 1600, 1600, ""],
    ['Personalnebenkosten (%)', 0.40, "", "", "", "", "Normalerweise 40%"],
    ['--- BERECHNETE PERSONALKOSTEN (MEUR) ---', '', '', '', '', '', ''],
    ['Löhne/Gehälter', "=(B10*B11)/1000000", "=(C10*C11)/1000000", "=(D10*D11)/1000000", "=(E10*E11)/1000000", "=(F10*F11)/1000000", "=SUM(B17:F17)"],
    ['Einstell./Entlass./Training', "=(B7*B12 + B8*B13 + B10*B14)/1000000", "=(C7*C12 + C8*C13 + C10*C14)/1000000", "=(D7*D12 + D8*D13 + D10*D14)/1000000", "=(E7*E12 + E8*E13 + E10*E14)/1000000", "=(F7*F12 + F8*F13 + F10*F14)/1000000", "=SUM(B18:F18)"],
    ['Personalnebenkosten', "=B17*$B$15", "=C17*$B$15", "=D17*$B$15", "=E17*$B$15", "=F17*$B$15", "=SUM(B19:F19)"],
    ['SUMME Personalkosten', "=B17+B18+B19", "=C17+C18+C19", "=D17+D18+D19", "=E17+E18+E19", "=F17+F18+F19", "=SUM(B20:F20)"],
    ['--- FERTIGUNG (SACHKOSTEN) ---', '', '', ''],
    ['Material- & Maschinenkosten', 50.0, 'MEUR', 'Reine Sachkosten Fertigung'],
    ['--- WEITERE OPERATIVE KOSTEN ---', '', '', ''],
    ['Total Verwaltungskosten', 10.0, 'MEUR', ''],
    ['Total Marketing/Vertrieb', 15.0, 'MEUR', ''],
    ['Forschung & Entwicklung (F&E)', 2.50, 'MEUR', 'Ökologie, Technologie, etc.'],
    ['Marktforschung', 0.15, 'MEUR', 'Kosten für Berichte / Analysen'],
    ['--- INVESTITIONEN ---', '', '', ''],
    ['Abschreibungen Alt-Anlagen', 6.0, 'MEUR', 'inklusive 0.25 MEUR fixe Gebäude-AfA'],
    ['Geplante Neuinvestition (Fertigung)', 15.0, 'MEUR', 'Totaler Kaufpreis Anlage - Cash Out!'],
    ['Geplante Neuinvestition (Umwelt)', 0.0, 'MEUR', 'Investition in Umwelttechnik - Cash Out!'],
    ['Geplante Neu-Abschreibung', 1.5, 'MEUR', 'Z.B. 10% Lineare Abschreibung'],
    ['Desinvestition/Verkauf Anlagen', 0.0, 'MEUR', 'Einzahlung bei Anlagenverkauf'],
    ['Veränderung Pensionsrückstellungen', 2.0, 'MEUR', 'Positiv = Cash-plus (Nicht zahlungswirksam!)']
]
for row in data2_base: ws2.append(row)

# =========================================================
# BLATT 3: SZENARIEN
# =========================================================
ws3 = wb.create_sheet('3_Szenario_Rechner')
data3 = [
    ['Feld', 'Szenario A (Defensiv)', 'Szenario B (Aggressiv)', 'Szenario C'],
    ['--- 1: EINGABEN (DEINE CFO ENTSCHEIDUNG) ---', '', '', ''],
    ['NEUER Kurzfristiger Kredit (MEUR)', 15.0, 20.0, 10.0],
    ['NEUER Langfristiger Kredit (MEUR)', 0.0, 10.0, 0.0],
    ['Dividende (MEUR)', 1.0, 0.0, 2.0],
    ['Geplanter Wertpapierkauf', 0.0, 0.0, 0.0]
]
for row in data3: ws3.append(row)

def add_scenario_row(ws, label, formula_template):
    row = [label]
    for col in ['B', 'C', 'D']:
        row.append(formula_template.replace('{COL}', col))
    ws.append(row)

ws3.append(['--- 2: BERECHNUNG EBIT & GEWINN ---', '', '', ''])
add_scenario_row(ws3, 'Plan-Umsatz', "='2_Team_Inputs'!B2*'2_Team_Inputs'!B3/1000000")
add_scenario_row(ws3, 'Plan-Bestandsveränderung', "=IF('2_Team_Inputs'!B4>0, ('2_Team_Inputs'!B4-'2_Team_Inputs'!B2)*( ('2_Team_Inputs'!$B$22+'2_Team_Inputs'!$D$20)/'2_Team_Inputs'!B4 ), 0)")
add_scenario_row(ws3, 'Plan-EBIT', "={COL}8+{COL}9-'2_Team_Inputs'!$G$20-'2_Team_Inputs'!$B$22-SUM('2_Team_Inputs'!$B$24:$B$27)-'2_Team_Inputs'!$B$29-'2_Team_Inputs'!$B$32")
add_scenario_row(ws3, 'Plan-Zinsaufwand (Kredite)', "=({COL}3+'1_Ist_Daten'!$B$8)*('1_Ist_Daten'!$B$15+'1_Ist_Daten'!$B$18) + ({COL}4*('1_Ist_Daten'!$B$14+'1_Ist_Daten'!$B$18)) + ('1_Ist_Daten'!$B$10*'1_Ist_Daten'!$B$14) + ('1_Ist_Daten'!$B$9*'1_Ist_Daten'!$B$16)")
add_scenario_row(ws3, 'Zinsertrag Wertpapiere', "={COL}6*'1_Ist_Daten'!$B$12")
add_scenario_row(ws3, 'Plan-EBT (Ergebnis vor Steuern)', "={COL}10-{COL}11+{COL}12")
add_scenario_row(ws3, 'Bemessungsgrundlage Steuer', "=MAX(0, {COL}13-'1_Ist_Daten'!$B$6)")
add_scenario_row(ws3, 'Plan-Steuern', "={COL}14*'1_Ist_Daten'!$B$13")
add_scenario_row(ws3, 'Plan-Jahresüberschuss', "={COL}13-{COL}15")

ws3.append(['--- 3: TOPSIM ZIELEINGABEN ---', '', '', ''])
add_scenario_row(ws3, 'Plan-Eigenkapital', "='1_Ist_Daten'!$B$4+{COL}16-{COL}5")
add_scenario_row(ws3, 'Plan-EKR (%)', "=({COL}16/{COL}18)*100")
add_scenario_row(ws3, 'Plan-Operativer Cashflow (MEUR)', "={COL}16+'2_Team_Inputs'!$B$29+'2_Team_Inputs'!$B$32+'2_Team_Inputs'!$B$34")

ws3.append(['--- 4: ZUR LIQUIDITÄTSPLANUNG ---', '', '', ''])
add_scenario_row(ws3, 'Kassenbestand ENDE', "='4_Liquiditaetsplanung'!{COL}24")
add_scenario_row(ws3, 'Warnung', '=IF({COL}22<0.1, "🔴 OVERDRAFT", "✅ OK")')


# =========================================================
# BLATT 4: LIQUIDITÄTSPLANUNG
# =========================================================
ws4 = wb.create_sheet('4_Liquiditaetsplanung')
ws4.append(['Feld', 'Szenario A (Defensiv)', 'Szenario B (Aggressiv)', 'Szenario C'])
ws4.append(['--- EINZAHLUNGEN (MEUR) ---', '', '', ''])
add_scenario_row(ws4, 'Einzahlungen aus Umsatz lfd. Periode', "='3_Szenario_Rechner'!{COL}8 * 0.80")
add_scenario_row(ws4, 'Einzahlungen Forderungen Vorperiode', "='1_Ist_Daten'!$B$3")
add_scenario_row(ws4, 'Rückzahlung Wertpapiere Vorperiode', "='1_Ist_Daten'!$B$11")
add_scenario_row(ws4, 'Zinsertrag Wertpapiere (Auszahlung hier)', "='3_Szenario_Rechner'!{COL}12")
add_scenario_row(ws4, 'Kreditaufnahme Kurzfristig', "='3_Szenario_Rechner'!{COL}3")
add_scenario_row(ws4, 'Kreditaufnahme Langfristig', "='3_Szenario_Rechner'!{COL}4")
add_scenario_row(ws4, 'Desinvestition/Verkauf Anlagen', "='2_Team_Inputs'!$B$33")
add_scenario_row(ws4, 'SUMME EINZAHLUNGEN', "=SUM({COL}3:{COL}9)")

ws4.append(['--- AUSZAHLUNGEN (MEUR) ---', '', '', ''])
add_scenario_row(ws4, 'Material- & Maschinenkosten', "='2_Team_Inputs'!$B$22")
add_scenario_row(ws4, 'Personalkosten (Total)', "='2_Team_Inputs'!$G$20")
add_scenario_row(ws4, 'Sonstige Aufwendungen (Budgets)', "=SUM('2_Team_Inputs'!$B$24:$B$27)")
add_scenario_row(ws4, 'Rückzahlung Alt-Kredite (Kurz + Überziehung)', "='1_Ist_Daten'!$B$8 + '1_Ist_Daten'!$B$9")
add_scenario_row(ws4, 'Plan-Neuinvestitionen (Fertigung + Umwelt)', "='2_Team_Inputs'!$B$30 + '2_Team_Inputs'!$B$31")
add_scenario_row(ws4, 'Wertpapierkauf', "='3_Szenario_Rechner'!{COL}6")
add_scenario_row(ws4, 'Auszahlung Dividende', "='3_Szenario_Rechner'!{COL}5")
add_scenario_row(ws4, 'Zinsaufwand (Kredite)', "='3_Szenario_Rechner'!{COL}11")
add_scenario_row(ws4, 'Steuern', "='3_Szenario_Rechner'!{COL}15")
add_scenario_row(ws4, 'SUMME AUSZAHLUNGEN', "=SUM({COL}12:{COL}20)")

ws4.append(['--- KASSENBESTIMMUNG (MEUR) ---', '', '', ''])
add_scenario_row(ws4, 'Kassenanfangsbestand', "='1_Ist_Daten'!$B$2")
add_scenario_row(ws4, 'PLAN-KASSENENDBESTAND', "={COL}23 + {COL}10 - {COL}21")

# =========================================================
# STYLING
# =========================================================
def apply_bold_headers(ws):
    for col in range(1, 8):
        try:
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
        except:
            pass
    ws.column_dimensions['A'].width = 45

for ws in [ws1, ws2, ws3, ws4]:
    apply_bold_headers(ws)
    for row in ws.iter_rows():
        if row[0].value and isinstance(row[0].value, str) and '---' in row[0].value:
            row[0].font = Font(bold=True)

# Spezifisches Styling Blatt 2 (Prozentzellen)
ws2.cell(row=15, column=2).number_format = '0%'

# Bedingte Formatierung (Warnungen)
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
white_font = Font(color='FFFFFF', bold=True)
black_font = Font(color='000000', bold=True)

# Warnung Kasse < 0
ws3.conditional_formatting.add('B22:D22', CellIsRule(operator='lessThan', formula=['0.1'], stopIfTrue=True, fill=red_fill, font=white_font))
ws4.conditional_formatting.add('B24:D24', CellIsRule(operator='lessThan', formula=['0.1'], stopIfTrue=True, fill=red_fill, font=white_font))

# Warnung Dividende > Gewinnvortrag
ws3.conditional_formatting.add('B5:D5', CellIsRule(operator='greaterThan', formula=["'1_Ist_Daten'!$B$5"], stopIfTrue=True, fill=yellow_fill, font=black_font))

# =========================================================
# SPEICHERN
# =========================================================
file_path = os.path.join(os.getcwd(), 'TOPSIM_CFO_Dashboard.xlsx')
if os.path.exists(file_path):
    backup_path = os.path.join(os.getcwd(), f"TOPSIM_CFO_Dashboard_Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy(file_path, backup_path)
    print(f"Alte Version gesichert unter: {backup_path}")

wb.save(file_path)
print(f"Neues professionelles Excel-Dashboard erstellt: {file_path}")