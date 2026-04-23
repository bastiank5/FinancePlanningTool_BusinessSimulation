import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

path = r"c:\Users\Bastian\OneDrive\Hochschule\Sechstes Semester\Business Simulation\Finance Tool\TOPSIM_Planungstool_V8.xlsx"
try:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
except Exception as e:
    print(f"ERROR: {e}")
    sys.exit(1)

print("Sheets:", wb.sheetnames)
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    row_count = 0
    for row in sheet.iter_rows(max_row=200, values_only=False):
        vals = []
        for cell in row:
            v = cell.value
            if v is None:
                v = ""
            vals.append(str(v))
        joined = " | ".join(vals)
        if any(v.strip() for v in vals):
            print(f"  Row {row_count}: {joined}")
        row_count += 1
wb.close()
