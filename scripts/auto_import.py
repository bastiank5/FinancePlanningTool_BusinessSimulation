import os
import glob
import sys
import subprocess

# Auto-install pip packages falls nötig
try:
    import xlrd
    import openpyxl
except ImportError:
    print("Installiere benötigte Bibliotheken (xlrd, openpyxl)...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "xlrd", "openpyxl"])
    import xlrd
    import openpyxl

base_dir = r"c:\Users\Bastian\OneDrive\Hochschule\Sechstes Semester\Business Simulation\Finance Tool"
sources_dir = os.path.join(base_dir, "Sources")

print("--- TOPSIM Auto-Importer ---")

# Finde den Ordner mit der höchsten Periode
period_folders = [f for f in os.listdir(sources_dir) if f.startswith("Periode ") and os.path.isdir(os.path.join(sources_dir, f))]
highest_period = -1
for f in period_folders:
    try:
        num = int(f.split(" ")[1])
        if num > highest_period:
            highest_period = num
    except:
        pass

if highest_period == -1:
    print("FEHLER: Keine Periode-Ordner (z.B. 'Periode 2') gefunden!")
    sys.exit(1)

latest_folder = os.path.join(sources_dir, f"Periode {highest_period}")
xls_files = glob.glob(os.path.join(latest_folder, "*.xls"))

if not xls_files:
    print(f"FEHLER: Keine .xls Datei in {latest_folder} gefunden!")
    sys.exit(1)

target_xls = xls_files[0]
print(f"Lese Daten aus aktuellster Datei: {target_xls}")

try:
    wb_source = xlrd.open_workbook(target_xls)
except Exception as e:
    print(f"FEHLER beim Öffnen der XLS: {e}")
    sys.exit(1)

# Extraktion der Kennzahlen (für U3 / Periode Aktuell)
kasse = 0.0
ek = 0.0
vortrag = 0.0
pension = 0.0

try:
    sheet_ex = wb_source.sheet_by_name("1) Executive Summary")
    for r in range(sheet_ex.nrows):
        lbl = str(sheet_ex.cell_value(r, 0)).strip()
        if "Kassenendbestand" in lbl:
            kasse = float(sheet_ex.cell_value(r, 3))
        if "Eigenkapital" == lbl:
            ek = float(sheet_ex.cell_value(r, 3))
except Exception as e:
    print(f"Warnung: Fehler beim Lesen vom Executive Summary: {e}")

try:
    sheet_bil = wb_source.sheet_by_name("14) Bilanz")
    for r in range(sheet_bil.nrows):
        lbl_passiva = str(sheet_bil.cell_value(r, 3)).strip()
        if "Pensionsrückstellungen" in lbl_passiva:
            pension = float(sheet_bil.cell_value(r, 4))
        if "Gewinn-/Verlustvortrag" in lbl_passiva:
            vortrag = float(sheet_bil.cell_value(r, 4))
except Exception as e:
    print(f"Warnung: Fehler beim Lesen der Bilanz: {e}")

print(f"✓ Extrahiert: Kasse={kasse} MEUR, EK={ek} MEUR, Vortrag={vortrag} MEUR, Pensionsrückst.={pension} MEUR")

# Update Excel Dashboard
db_path = os.path.join(base_dir, "TOPSIM_CFO_Dashboard.xlsx")
if not os.path.exists(db_path):
    print("FEHLER: TOPSIM_CFO_Dashboard.xlsx nicht gefunden.")
    sys.exit(1)

try:
    wb_db = openpyxl.load_workbook(db_path)
    ws1 = wb_db['1_Ist_Daten']
    
    # Prüfe ob Periode X bereits existiert, ansonsten hänge neue Spalte an
    col_to_write = ws1.max_column + 1
    for c in range(1, ws1.max_column + 1):
        if str(ws1.cell(row=1, column=c).value) == f"Periode {highest_period}":
            col_to_write = c
            break

    # Schreibe Werte
    ws1.cell(row=1, column=col_to_write, value=f"Periode {highest_period}")
    from openpyxl.styles import Font
    ws1.cell(row=1, column=col_to_write).font = Font(bold=True)
    ws1.cell(row=2, column=col_to_write, value=kasse)
    ws1.cell(row=3, column=col_to_write, value=ek)
    ws1.cell(row=4, column=col_to_write, value=vortrag)
    ws1.cell(row=5, column=col_to_write, value=pension)
    # Zinsen & Steuern voreingestellt übernehmen, falls der Nutzer sie nicht im Handbuch ändert
    ws1.cell(row=6, column=col_to_write, value=0.30)
    ws1.cell(row=7, column=col_to_write, value=0.06)
    ws1.cell(row=8, column=col_to_write, value=0.08)

    wb_db.save(db_path)
    print(f"✓ ERFOLG! Periode {highest_period} wurde als neue Spalte ins CFO-Dashboard eingetragen.")
except Exception as e:
    print(f"Fehler beim Speichern der Excel: {e}")
