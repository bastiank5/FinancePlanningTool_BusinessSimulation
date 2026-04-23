"""Read reports-2.xls and TOPSIM_Planungstool_V8.xlsx and dump all sheet contents."""
import xlrd
import openpyxl
import sys

def read_xls(filepath):
    """Read .xls (old format) using xlrd."""
    print(f"\n{'='*80}")
    print(f"FILE: {filepath}")
    print(f"{'='*80}")
    wb = xlrd.open_workbook(filepath)
    for sheet_name in wb.sheet_names():
        sheet = wb.sheet_by_name(sheet_name)
        print(f"\n--- Sheet: {sheet_name} ({sheet.nrows} rows x {sheet.ncols} cols) ---")
        for row_idx in range(min(sheet.nrows, 100)):
            row = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    val = ""
                row.append(str(val))
            # Only print rows with some content
            if any(v.strip() for v in row):
                print(f"  Row {row_idx}: {' | '.join(row)}")

def read_xlsx(filepath):
    """Read .xlsx using openpyxl."""
    print(f"\n{'='*80}")
    print(f"FILE: {filepath}")
    print(f"{'='*80}")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        print(f"Error opening: {e}")
        return
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        row_count = 0
        for row in sheet.iter_rows(max_row=150, values_only=False):
            vals = []
            for cell in row:
                v = cell.value
                if v is None:
                    v = ""
                vals.append(str(v))
            if any(v.strip() for v in vals):
                print(f"  Row {row_count}: {' | '.join(vals)}")
            row_count += 1
    wb.close()

if __name__ == "__main__":
    import os
    base = r"c:\Users\Bastian\OneDrive\Hochschule\Sechstes Semester\Business Simulation\Finance Tool"
    
    # Read Period 2 results
    reports_path = os.path.join(base, "Sources", "Periode 2", "reports-2.xls")
    if os.path.exists(reports_path):
        read_xls(reports_path)
    
    # Read Planning tool
    plan_path = os.path.join(base, "TOPSIM_Planungstool_V8.xlsx")
    if os.path.exists(plan_path):
        read_xlsx(plan_path)
